from fastapi import APIRouter, HTTPException, Query, UploadFile, File, BackgroundTasks, Response
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from sqlalchemy import select, desc, func, and_, or_
from sqlalchemy.orm import selectinload
import json
import csv
import io
import zipfile
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import tempfile

from database import get_session, DatabaseManager
from models import Video, Task, Channel, SearchQuery, Analytics
from tasks import (
    parse_youtube_search,
    parse_youtube_channel,
    analyze_video_content,
    pause_task as celery_pause_task,
    resume_task as celery_resume_task,
    cancel_task as celery_cancel_task
)
from websocket import notify_task_update, notify_video_added, send_notification

# Импорт логгера
import logging

logger = logging.getLogger(__name__)

# Создание роутера
api_router = APIRouter()

# ==================== Videos API ====================

@api_router.get("/videos")
async def get_videos(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    search: Optional[str] = None,
    video_type: Optional[str] = None,
    sort: str = Query("date", regex="^(date|views|engagement)$")
):
    """Получение списка видео с фильтрацией и пагинацией"""
    async with get_session() as session:
        # Базовый запрос
        query = select(Video).where(Video.parse_status == "completed")
        
        # Фильтрация по поиску
        if search:
            search_filter = or_(
                Video.title.ilike(f"%{search}%"),
                Video.description.ilike(f"%{search}%")
            )
            query = query.where(search_filter)
        
        # Фильтрация по типу видео
        if video_type == "shorts":
            query = query.where(Video.is_short == True)
        elif video_type == "long":
            query = query.where(Video.is_short == False)
        
        # Сортировка
        if sort == "views":
            query = query.order_by(desc(Video.views))
        elif sort == "engagement":
            query = query.order_by(desc(Video.engagement_rate))
        else:
            query = query.order_by(desc(Video.publish_date))
        
        # Подсчет общего количества
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar() or 0
        
        # Пагинация
        offset = (page - 1) * per_page
        query = query.limit(per_page).offset(offset)
        
        # Выполнение запроса
        result = await session.execute(query)
        videos = result.scalars().all()
        
        # Формирование ответа с ВСЕМИ полями
        return {
            "videos": [
                {
                    "id": video.id,
                    "video_url": video.video_url,
                    "title": video.title,
                    "description": video.description,
                    "channel_url": video.channel_url,
                    "channel_name": video.channel_url.split("/")[-1] if video.channel_url else "Unknown",
                    "views": video.views,
                    "likes": video.likes,
                    "comments": video.comments,
                    "dislikes": video.dislikes or 0,
                    "engagement_rate": round(video.engagement_rate, 2),
                    "like_ratio": round(video.like_ratio, 2),
                    "comment_ratio": round(video.comment_ratio, 2),
                    "publish_date": video.publish_date.isoformat() if video.publish_date else None,
                    "is_short": video.is_short,
                    "thumbnail_url": video.thumbnail_url,
                    "duration": video.duration,
                    "video_category": video.video_category,
                    "video_quality": video.video_quality,
                    "has_cc": video.has_cc,
                    "has_chapters": video.has_chapters,
                    "has_branding": video.has_branding,
                    "has_intro": video.has_intro,
                    "has_outro": video.has_outro,
                    "emoji_in_title": video.emoji_in_title,
                    "top_5_keywords": video.top_5_keywords or [],
                    "links_in_description": video.links_in_description or 0,
                    "has_pinned_comment": video.has_pinned_comment,
                    "speech_speed": video.speech_speed,
                    "improvement_recommendations": video.improvement_recommendations,
                    "success_analysis": video.success_analysis,
                    "content_strategy": video.content_strategy,
                    "channel_avg_views": video.channel_avg_views,
                    "channel_avg_likes": video.channel_avg_likes,
                    "channel_frequency": video.channel_frequency,
                    "channel_age": video.channel_age,
                    "subtitles": video.subtitles,
                    "keywords": video.keywords
                }
                for video in videos
            ],
            "pagination": {
                "total": total,
                "page": page,
                "per_page": per_page,
                "pages": (total + per_page - 1) // per_page if total > 0 else 0
            }
        }

@api_router.get("/videos/{video_id}")
async def get_video(video_id: int):
    """Получение детальной информации о видео"""
    async with get_session() as session:
        stmt = select(Video).where(Video.id == video_id)
        result = await session.execute(stmt)
        video = result.scalar_one_or_none()
        
        if not video:
            raise HTTPException(status_code=404, detail="Видео не найдено")
        
        # Получаем информацию о канале
        channel_info = None
        if video.channel_url:
            channel_stmt = select(Channel).where(
                Channel.channel_url == video.channel_url
            )
            channel_result = await session.execute(channel_stmt)
            channel = channel_result.scalar_one_or_none()
            if channel:
                channel_info = {
                    "title": channel.title,
                    "subscribers": channel.subscriber_count,
                    "videos": channel.video_count,
                    "avg_views": channel.avg_views
                }
        
        # Возвращаем ВСЕ поля
        return {
            "id": video.id,
            "video_url": video.video_url,
            "title": video.title,
            "description": video.description,
            "channel_url": video.channel_url,
            "channel_info": channel_info,
            "is_short": video.is_short,
            "views": video.views,
            "likes": video.likes,
            "comments": video.comments,
            "dislikes": video.dislikes,
            "like_ratio": video.like_ratio,
            "comment_ratio": video.comment_ratio,
            "engagement_rate": video.engagement_rate,
            "duration": video.duration,
            "publish_date": video.publish_date.isoformat() if video.publish_date else None,
            "thumbnail_url": video.thumbnail_url,
            "video_category": video.video_category,
            "video_quality": video.video_quality,
            "has_cc": video.has_cc,
            "has_chapters": video.has_chapters,
            "has_branding": video.has_branding,
            "has_intro": video.has_intro,
            "has_outro": video.has_outro,
            "emoji_in_title": video.emoji_in_title,
            "keywords": video.keywords,
            "top_5_keywords": video.top_5_keywords,
            "subtitles": video.subtitles,
            "links_in_description": video.links_in_description,
            "links_in_channel_description": video.links_in_channel_description,
            "has_pinned_comment": video.has_pinned_comment,
            "contacts_in_video": video.contacts_in_video,
            "contacts_in_channel": video.contacts_in_channel,
            "speech_speed": video.speech_speed,
            "channel_avg_views": video.channel_avg_views,
            "channel_avg_likes": video.channel_avg_likes,
            "channel_frequency": video.channel_frequency,
            "channel_age": video.channel_age,
            "average_view_duration": video.average_view_duration,
            "click_through_rate": video.click_through_rate,
            "improvement_recommendations": video.improvement_recommendations,
            "success_analysis": video.success_analysis,
            "content_strategy": video.content_strategy,
            "created_at": video.created_at.isoformat() if video.created_at else None
        }

@api_router.post("/videos/{video_id}/analyze")
async def analyze_video(video_id: int):
    """Запуск глубокого анализа видео"""
    async with get_session() as session:
        # Проверка существования видео
        stmt = select(Video).where(Video.id == video_id)
        result = await session.execute(stmt)
        video = result.scalar_one_or_none()
        
        if not video:
            raise HTTPException(status_code=404, detail="Видео не найдено")
        
        # Создание задачи анализа
        task = Task(
            task_id=f"analysis_{video_id}_{datetime.now().timestamp()}",
            task_type="analysis",
            parameters={"video_id": video_id, "video_title": video.title},
            status="pending"
        )
        session.add(task)
        await session.commit()
        
        # Запуск Celery задачи
        celery_task = analyze_video_content.delay(task.id, video_id)
        task.task_id = celery_task.id
        await session.commit()
        
        # Уведомление через WebSocket
        await notify_task_update(task.to_dict())
        
        return {"task_id": task.id, "status": "started", "celery_id": celery_task.id}

# ИСПРАВЛЕННЫЙ ЭКСПОРТ - теперь возвращает файл напрямую
@api_router.get("/videos/export/{format}")
async def export_videos(
    format: str,
    ids: Optional[str] = None
):
    """Экспорт видео в CSV, JSON или Excel"""
    if format not in ["csv", "json", "excel"]:
        raise HTTPException(status_code=400, detail="Неверный формат. Используйте: csv, json или excel")
    
    async with get_session() as session:
        # Базовый запрос
        stmt = select(Video).where(Video.parse_status == "completed")
        
        # Если переданы конкретные ID видео (от фильтров)
        if ids:
            video_ids = [int(id) for id in ids.split(',') if id.isdigit()]
            if video_ids:
                stmt = stmt.where(Video.id.in_(video_ids))
        
        # Сортировка по просмотрам
        stmt = stmt.order_by(desc(Video.views))
        
        result = await session.execute(stmt)
        videos = result.scalars().all()
        
        if not videos:
            raise HTTPException(status_code=404, detail="Нет данных для экспорта")
        
        # Создаем временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False)
        
        try:
            if format == "csv":
                # CSV экспорт
                with open(temp_file.name, 'w', encoding='utf-8-sig', newline='') as f:
                    fieldnames = [
                        "video_url", "channel_url", "title", "description", "is_short", "views", "likes", 
                        "comments", "dislikes", "like_ratio", "comment_ratio", "engagement_rate", 
                        "publish_date", "duration", "video_category", "video_quality", "has_cc", 
                        "has_chapters", "has_branding", "has_intro", "has_outro", "emoji_in_title", 
                        "top_5_keywords", "links_in_description", "links_in_channel_description", 
                        "has_pinned_comment", "contacts_in_video", "contacts_in_channel", 
                        "speech_speed", "channel_avg_views", "channel_avg_likes", "channel_frequency", 
                        "channel_age", "average_view_duration", "click_through_rate",
                        "improvement_recommendations", "success_analysis", "content_strategy",
                        "subtitles", "keywords"
                    ]
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for video in videos:
                        writer.writerow({
                            "video_url": video.video_url,
                            "channel_url": video.channel_url,
                            "title": video.title,
                            "description": video.description or "",
                            "is_short": "Да" if video.is_short else "Нет",
                            "views": video.views,
                            "likes": video.likes,
                            "comments": video.comments,
                            "dislikes": video.dislikes or 0,
                            "like_ratio": f"{video.like_ratio:.2f}%",
                            "comment_ratio": f"{video.comment_ratio:.2f}%",
                            "engagement_rate": f"{video.engagement_rate:.2f}%",
                            "publish_date": video.publish_date.strftime("%Y-%m-%d %H:%M:%S") if video.publish_date else "",
                            "duration": video.duration or "",
                            "video_category": video.video_category or "",
                            "video_quality": video.video_quality or "",
                            "has_cc": "Да" if video.has_cc else "Нет",
                            "has_chapters": "Да" if video.has_chapters else "Нет",
                            "has_branding": "Да" if video.has_branding else "Нет",
                            "has_intro": "Да" if video.has_intro else "Нет",
                            "has_outro": "Да" if video.has_outro else "Нет",
                            "emoji_in_title": "Да" if video.emoji_in_title else "Нет",
                            "top_5_keywords": ", ".join(video.top_5_keywords or []),
                            "links_in_description": video.links_in_description or 0,
                            "links_in_channel_description": video.links_in_channel_description or 0,
                            "has_pinned_comment": "Да" if video.has_pinned_comment else "Нет",
                            "contacts_in_video": json.dumps(video.contacts_in_video or [], ensure_ascii=False),
                            "contacts_in_channel": json.dumps(video.contacts_in_channel or [], ensure_ascii=False),
                            "speech_speed": video.speech_speed or 0,
                            "channel_avg_views": video.channel_avg_views or 0,
                            "channel_avg_likes": video.channel_avg_likes or 0,
                            "channel_frequency": video.channel_frequency or 0,
                            "channel_age": video.channel_age or 0,
                            "average_view_duration": video.average_view_duration or 0,
                            "click_through_rate": video.click_through_rate or 0,
                            "improvement_recommendations": video.improvement_recommendations or "",
                            "success_analysis": video.success_analysis or "",
                            "content_strategy": video.content_strategy or "",
                            "subtitles": video.subtitles or "",
                            "keywords": json.dumps(video.keywords or [], ensure_ascii=False)
                        })
                
                return FileResponse(
                    temp_file.name,
                    media_type="text/csv; charset=utf-8",
                    filename=f"youtube_videos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                )
                
            elif format == "json":
                # JSON экспорт
                data = []
                for video in videos:
                    data.append({
                        "video_url": video.video_url,
                        "title": video.title,
                        "description": video.description,
                        "channel_url": video.channel_url,
                        "is_short": video.is_short,
                        "metrics": {
                            "views": video.views,
                            "likes": video.likes,
                            "comments": video.comments,
                            "dislikes": video.dislikes or 0,
                            "engagement_rate": video.engagement_rate,
                            "like_ratio": video.like_ratio,
                            "comment_ratio": video.comment_ratio
                        },
                        "content": {
                            "duration": video.duration,
                            "category": video.video_category,
                            "quality": video.video_quality,
                            "has_cc": video.has_cc,
                            "has_chapters": video.has_chapters,
                            "keywords": video.keywords,
                            "top_5_keywords": video.top_5_keywords,
                            "subtitles": video.subtitles,
                            "links_in_description": video.links_in_description
                        },
                        "channel_stats": {
                            "avg_views": video.channel_avg_views,
                            "avg_likes": video.channel_avg_likes,
                            "frequency": video.channel_frequency,
                            "age": video.channel_age
                        },
                        "analysis": {
                            "recommendations": video.improvement_recommendations,
                            "success_factors": video.success_analysis,
                            "strategy": video.content_strategy
                        },
                        "publish_date": video.publish_date.isoformat() if video.publish_date else None,
                        "created_at": video.created_at.isoformat() if video.created_at else None
                    })
                
                with open(temp_file.name, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                return FileResponse(
                    temp_file.name,
                    media_type="application/json; charset=utf-8",
                    filename=f"youtube_videos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                )
                
            else:  # Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "YouTube Видео"
                
                # Заголовки - ВСЕ 42+ параметра
                headers = [
                    "URL видео", "URL канала", "Название", "Описание", "Тип видео", "Просмотры", 
                    "Лайки", "Комментарии", "Дизлайки", "% лайков", "% комментариев", 
                    "Вовлеченность %", "Дата публикации", "Длительность", "Категория", 
                    "Качество", "Субтитры", "Главы", "Брендинг", "Интро", "Аутро", 
                    "Эмодзи в заголовке", "Топ-5 ключевых слов", "Ссылок в описании", 
                    "Ссылок в канале", "Закрепленный комментарий", "Контакты в видео", 
                    "Контакты в канале", "Скорость речи", "Средние просмотры канала", 
                    "Средние лайки канала", "Частота публикаций", "Возраст канала", 
                    "Средняя длительность просмотра", "CTR", "Рекомендации", 
                    "Анализ успеха", "Стратегия", "Транскрипт", "Ключевые слова"
                ]
                
                # Стили для заголовков
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Данные
                for row, video in enumerate(videos, 2):
                    ws.cell(row=row, column=1, value=video.video_url)
                    ws.cell(row=row, column=2, value=video.channel_url)
                    ws.cell(row=row, column=3, value=video.title)
                    ws.cell(row=row, column=4, value=video.description or "")
                    ws.cell(row=row, column=5, value="Short" if video.is_short else "Обычное")
                    ws.cell(row=row, column=6, value=video.views)
                    ws.cell(row=row, column=7, value=video.likes)
                    ws.cell(row=row, column=8, value=video.comments)
                    ws.cell(row=row, column=9, value=video.dislikes or 0)
                    ws.cell(row=row, column=10, value=f"{video.like_ratio:.2f}%")
                    ws.cell(row=row, column=11, value=f"{video.comment_ratio:.2f}%")
                    ws.cell(row=row, column=12, value=f"{video.engagement_rate:.2f}%")
                    ws.cell(row=row, column=13, value=video.publish_date.strftime("%Y-%m-%d %H:%M:%S") if video.publish_date else "")
                    ws.cell(row=row, column=14, value=video.duration or "")
                    ws.cell(row=row, column=15, value=video.video_category or "")
                    ws.cell(row=row, column=16, value=video.video_quality or "")
                    ws.cell(row=row, column=17, value="Да" if video.has_cc else "Нет")
                    ws.cell(row=row, column=18, value="Да" if video.has_chapters else "Нет")
                    ws.cell(row=row, column=19, value="Да" if video.has_branding else "Нет")
                    ws.cell(row=row, column=20, value="Да" if video.has_intro else "Нет")
                    ws.cell(row=row, column=21, value="Да" if video.has_outro else "Нет")
                    ws.cell(row=row, column=22, value="Да" if video.emoji_in_title else "Нет")
                    ws.cell(row=row, column=23, value=", ".join(video.top_5_keywords or []))
                    ws.cell(row=row, column=24, value=video.links_in_description or 0)
                    ws.cell(row=row, column=25, value=video.links_in_channel_description or 0)
                    ws.cell(row=row, column=26, value="Да" if video.has_pinned_comment else "Нет")
                    ws.cell(row=row, column=27, value=json.dumps(video.contacts_in_video or [], ensure_ascii=False))
                    ws.cell(row=row, column=28, value=json.dumps(video.contacts_in_channel or [], ensure_ascii=False))
                    ws.cell(row=row, column=29, value=video.speech_speed or 0)
                    ws.cell(row=row, column=30, value=video.channel_avg_views or 0)
                    ws.cell(row=row, column=31, value=video.channel_avg_likes or 0)
                    ws.cell(row=row, column=32, value=video.channel_frequency or 0)
                    ws.cell(row=row, column=33, value=video.channel_age or 0)
                    ws.cell(row=row, column=34, value=video.average_view_duration or 0)
                    ws.cell(row=row, column=35, value=video.click_through_rate or 0)
                    ws.cell(row=row, column=36, value=video.improvement_recommendations or "")
                    ws.cell(row=row, column=37, value=video.success_analysis or "")
                    ws.cell(row=row, column=38, value=video.content_strategy or "")
                    ws.cell(row=row, column=39, value=video.subtitles[:500] + "..." if video.subtitles and len(video.subtitles) > 500 else video.subtitles or "")
                    ws.cell(row=row, column=40, value=json.dumps(video.keywords or [], ensure_ascii=False))
                
                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Закрепление заголовка
                ws.freeze_panes = "A2"
                
                # Добавляем второй лист со статистикой
                ws2 = wb.create_sheet("Статистика")
                ws2.append(["Параметр", "Значение"])
                ws2.append(["Всего видео", len(videos)])
                ws2.append(["Shorts", sum(1 for v in videos if v.is_short)])
                ws2.append(["Обычные видео", sum(1 for v in videos if not v.is_short)])
                ws2.append(["Средние просмотры", sum(v.views for v in videos) / len(videos) if videos else 0])
                ws2.append(["Средняя вовлеченность", sum(v.engagement_rate for v in videos) / len(videos) if videos else 0])
                
                # Сохранение в файл
                wb.save(temp_file.name)
                
                return FileResponse(
                    temp_file.name,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"youtube_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
        finally:
            # Удаляем временный файл после отправки
            def cleanup():
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
            
            import atexit
            atexit.register(cleanup)

# ==================== Channels API ====================

@api_router.get("/channels")
async def get_channels(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    search: Optional[str] = None,
    sort: str = Query("subscribers", regex="^(subscribers|videos|views|date)$")
):
    """Получение списка каналов с фильтрацией и пагинацией"""
    async with get_session() as session:
        # Базовый запрос
        query = select(Channel)
        
        # Фильтрация по поиску
        if search:
            search_filter = or_(
                Channel.title.ilike(f"%{search}%"),
                Channel.description.ilike(f"%{search}%")
            )
            query = query.where(search_filter)
        
        # Сортировка
        if sort == "subscribers":
            query = query.order_by(desc(Channel.subscriber_count))
        elif sort == "videos":
            query = query.order_by(desc(Channel.video_count))
        elif sort == "views":
            query = query.order_by(desc(Channel.view_count))
        else:
            query = query.order_by(desc(Channel.created_at))
        
        # Подсчет общего количества
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar() or 0
        
        # Пагинация
        offset = (page - 1) * per_page
        query = query.limit(per_page).offset(offset)
        
        # Выполнение запроса
        result = await session.execute(query)
        channels = result.scalars().all()
        
        return {
            "channels": [
                {
                    "id": channel.id,
                    "channel_id": channel.channel_id,
                    "channel_url": channel.channel_url,
                    "title": channel.title,
                    "description": channel.description[:200] + "..." if channel.description and len(channel.description) > 200 else channel.description,
                    "subscriber_count": channel.subscriber_count,
                    "video_count": channel.video_count,
                    "view_count": channel.view_count,
                    "avg_views": channel.avg_views,
                    "avg_likes": channel.avg_likes,
                    "avg_comments": channel.avg_comments,
                    "upload_frequency": channel.upload_frequency,
                    "country": channel.country,
                    "created_date": channel.created_date.isoformat() if channel.created_date else None,
                    "thumbnail_url": channel.thumbnail_url,
                    "last_parsed": channel.last_parsed.isoformat() if channel.last_parsed else None
                }
                for channel in channels
            ],
            "pagination": {
                "total": total,
                "page": page,
                "per_page": per_page,
                "pages": (total + per_page - 1) // per_page if total > 0 else 0
            }
        }

@api_router.get("/channels/{channel_id}")
async def get_channel(channel_id: int):
    """Получение детальной информации о канале"""
    async with get_session() as session:
        stmt = select(Channel).where(Channel.id == channel_id)
        result = await session.execute(stmt)
        channel = result.scalar_one_or_none()
        
        if not channel:
            raise HTTPException(status_code=404, detail="Канал не найден")
        
        # Получаем последние видео канала
        videos_stmt = select(Video).where(
            Video.channel_url == channel.channel_url
        ).order_by(desc(Video.publish_date)).limit(10)
        videos_result = await session.execute(videos_stmt)
        recent_videos = videos_result.scalars().all()
        
        return {
            "id": channel.id,
            "channel_id": channel.channel_id,
            "channel_url": channel.channel_url,
            "title": channel.title,
            "description": channel.description,
            "subscriber_count": channel.subscriber_count,
            "video_count": channel.video_count,
            "view_count": channel.view_count,
            "avg_views": channel.avg_views,
            "avg_likes": channel.avg_likes,
            "avg_comments": channel.avg_comments,
            "upload_frequency": channel.upload_frequency,
            "country": channel.country,
            "created_date": channel.created_date.isoformat() if channel.created_date else None,
            "thumbnail_url": channel.thumbnail_url,
            "keywords": channel.keywords,
            "content_categories": channel.content_categories,
            "target_audience": channel.target_audience,
            "last_parsed": channel.last_parsed.isoformat() if channel.last_parsed else None,
            "created_at": channel.created_at.isoformat() if channel.created_at else None,
            "recent_videos": [
                {
                    "id": video.id,
                    "title": video.title,
                    "views": video.views,
                    "publish_date": video.publish_date.isoformat() if video.publish_date else None,
                    "thumbnail_url": video.thumbnail_url
                }
                for video in recent_videos
            ]
        }

@api_router.get("/channels/export/{format}")
async def export_channels(format: str):
    """Экспорт каналов в различные форматы"""
    if format not in ["csv", "json", "excel"]:
        raise HTTPException(status_code=400, detail="Неверный формат. Используйте: csv, json или excel")
        
    async with get_session() as session:
        stmt = select(Channel).order_by(desc(Channel.subscriber_count))
        result = await session.execute(stmt)
        channels = result.scalars().all()
        
        if not channels:
            raise HTTPException(status_code=404, detail="Нет каналов для экспорта")
        
        temp_file = tempfile.NamedTemporaryFile(delete=False)
        
        try:
            if format == "csv":
                with open(temp_file.name, 'w', encoding='utf-8-sig', newline='') as f:
                    fieldnames = [
                        "channel_id", "channel_url", "title", "description", "subscriber_count",
                        "video_count", "view_count", "avg_views", "avg_likes", "avg_comments",
                        "upload_frequency", "country", "created_date", "keywords"
                    ]
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for channel in channels:
                        writer.writerow({
                            "channel_id": channel.channel_id,
                            "channel_url": channel.channel_url,
                            "title": channel.title,
                            "description": channel.description or "",
                            "subscriber_count": channel.subscriber_count,
                            "video_count": channel.video_count,
                            "view_count": channel.view_count,
                            "avg_views": int(channel.avg_views) if channel.avg_views else 0,
                            "avg_likes": int(channel.avg_likes) if channel.avg_likes else 0,
                            "avg_comments": int(channel.avg_comments) if channel.avg_comments else 0,
                            "upload_frequency": channel.upload_frequency or 0,
                            "country": channel.country or "",
                            "created_date": channel.created_date.strftime("%Y-%m-%d") if channel.created_date else "",
                            "keywords": json.dumps(channel.keywords or [], ensure_ascii=False)
                        })
                
                return FileResponse(
                    temp_file.name,
                    media_type="text/csv; charset=utf-8",
                    filename=f"youtube_channels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                )
                
            elif format == "json":
                data = []
                for channel in channels:
                    data.append({
                        "channel_id": channel.channel_id,
                        "channel_url": channel.channel_url,
                        "title": channel.title,
                        "description": channel.description,
                        "stats": {
                            "subscribers": channel.subscriber_count,
                            "videos": channel.video_count,
                            "total_views": channel.view_count,
                            "avg_views_per_video": channel.avg_views,
                            "avg_likes_per_video": channel.avg_likes,
                            "avg_comments_per_video": channel.avg_comments
                        },
                        "upload_frequency_days": channel.upload_frequency,
                        "country": channel.country,
                        "created_date": channel.created_date.isoformat() if channel.created_date else None,
                        "keywords": channel.keywords,
                        "last_parsed": channel.last_parsed.isoformat() if channel.last_parsed else None
                    })
                
                with open(temp_file.name, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                return FileResponse(
                    temp_file.name,
                    media_type="application/json; charset=utf-8",
                    filename=f"youtube_channels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                )
                
            else:  # Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "YouTube Каналы"
                
                headers = [
                    "ID канала", "URL канала", "Название", "Описание", "Подписчики",
                    "Количество видео", "Всего просмотров", "Средние просмотры",
                    "Средние лайки", "Средние комментарии", "Частота загрузки (дней)",
                    "Страна", "Дата создания", "Ключевые слова", "Последний парсинг"
                ]
                
                # Стили для заголовков
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Данные
                for row, channel in enumerate(channels, 2):
                    ws.cell(row=row, column=1, value=channel.channel_id)
                    ws.cell(row=row, column=2, value=channel.channel_url)
                    ws.cell(row=row, column=3, value=channel.title)
                    ws.cell(row=row, column=4, value=channel.description[:200] + "..." if channel.description and len(channel.description) > 200 else channel.description or "")
                    ws.cell(row=row, column=5, value=channel.subscriber_count)
                    ws.cell(row=row, column=6, value=channel.video_count)
                    ws.cell(row=row, column=7, value=channel.view_count)
                    ws.cell(row=row, column=8, value=int(channel.avg_views) if channel.avg_views else 0)
                    ws.cell(row=row, column=9, value=int(channel.avg_likes) if channel.avg_likes else 0)
                    ws.cell(row=row, column=10, value=int(channel.avg_comments) if channel.avg_comments else 0)
                    ws.cell(row=row, column=11, value=channel.upload_frequency or 0)
                    ws.cell(row=row, column=12, value=channel.country or "")
                    ws.cell(row=row, column=13, value=channel.created_date.strftime("%Y-%m-%d") if channel.created_date else "")
                    ws.cell(row=row, column=14, value=json.dumps(channel.keywords or [], ensure_ascii=False))
                    ws.cell(row=row, column=15, value=channel.last_parsed.strftime("%Y-%m-%d %H:%M:%S") if channel.last_parsed else "")
                
                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Закрепление заголовка
                ws.freeze_panes = "A2"
                
                # Сохранение
                wb.save(temp_file.name)
                
                return FileResponse(
                    temp_file.name,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"youtube_channels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
        finally:
            # Очистка
            def cleanup():
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
            
            import atexit
            atexit.register(cleanup)

# ==================== Tasks API ====================

@api_router.get("/tasks")
async def get_tasks():
    """Получение списка задач"""
    async with get_session() as session:
        stmt = select(Task).order_by(desc(Task.created_at))
        result = await session.execute(stmt)
        tasks = result.scalars().all()
        
        return [task.to_dict() for task in tasks]

@api_router.post("/parse")
async def create_parse_task(params: Dict[str, Any]):
    """Создание новой задачи парсинга"""
    async with get_session() as session:
        # Определение типа задачи
        if params.get("type") == "channel":
            task_type = "channel_parse"
        else:
            task_type = "search"
        
        # Создание задачи
        task = Task(
            task_id=f"{task_type}_{datetime.now().timestamp()}",
            task_type=task_type,
            parameters=params,
            status="pending"
        )
        session.add(task)
        await session.commit()
        
        # Запуск Celery задачи
        if task_type == "channel_parse":
            celery_task = parse_youtube_channel.delay(task.id, params)
        else:
            celery_task = parse_youtube_search.delay(task.id, params)
        
        task.task_id = celery_task.id
        await session.commit()
        
        # Уведомление через WebSocket
        await notify_task_update(task.to_dict())
        
        return task.to_dict()

@api_router.post("/tasks/{task_id}/pause")
async def pause_task(task_id: int):
    """Приостановка задачи"""
    async with get_session() as session:
        task = await session.get(Task, task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        
        if task.status != "running":
            raise HTTPException(status_code=400, detail="Можно приостановить только выполняющуюся задачу")
        
        # Приостановка Celery задачи
        celery_pause_task(task.task_id)
        
        # Обновление статуса
        task.status = "paused"
        task.paused_at = datetime.now()
        await session.commit()
        
        # Уведомление
        await notify_task_update(task.to_dict())
        
        return {"status": "paused", "task_id": task_id}

@api_router.post("/tasks/{task_id}/resume")
async def resume_task(task_id: int):
    """Возобновление задачи"""
    async with get_session() as session:
        task = await session.get(Task, task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        
        if task.status != "paused":
            raise HTTPException(status_code=400, detail="Можно возобновить только приостановленную задачу")
        
        # Возобновление Celery задачи
        celery_resume_task(task.task_id)
        
        # Обновление статуса
        task.status = "running"
        await session.commit()
        
        # Уведомление
        await notify_task_update(task.to_dict())
        
        return {"status": "resumed", "task_id": task_id}

@api_router.post("/tasks/{task_id}/cancel")
async def cancel_task(task_id: int):
    """Отмена задачи"""
    async with get_session() as session:
        task = await session.get(Task, task_id)
        if not task:
            raise HTTPException(status_code=404, detail="Задача не найдена")
        
        if task.status in ["completed", "failed", "cancelled"]:
            raise HTTPException(status_code=400, detail="Задача уже завершена")
        
        # Отмена Celery задачи
        celery_cancel_task(task.task_id)
        
        # Обновление статуса
        task.status = "cancelled"
        await session.commit()
        
        # Уведомление
        await notify_task_update(task.to_dict())
        
        return {"status": "cancelled", "task_id": task_id}

# ИСПРАВЛЕНИЕ: Приостановить все задачи
@api_router.post("/tasks/pause-all")
async def pause_all_tasks():
    """Приостановка всех активных задач"""
    async with get_session() as session:
        # Получаем все выполняющиеся задачи
        stmt = select(Task).where(Task.status == "running")
        result = await session.execute(stmt)
        running_tasks = result.scalars().all()
        
        paused_count = 0
        
        for task in running_tasks:
            try:
                # Приостанавливаем Celery задачу
                celery_pause_task(task.task_id)
                
                # Обновляем статус в БД
                task.status = "paused"
                task.paused_at = datetime.now()
                paused_count += 1
                
                # Уведомление через WebSocket
                await notify_task_update(task.to_dict())
                
            except Exception as e:
                logger.error(f"Failed to pause task {task.id}: {e}")
                continue
        
        await session.commit()
        
        return {
            "status": "success",
            "paused_count": paused_count,
            "message": f"Приостановлено задач: {paused_count}"
        }

# ИСПРАВЛЕНИЕ: Возобновить все задачи
@api_router.post("/tasks/resume-all")
async def resume_all_tasks():
    """Возобновление всех приостановленных задач"""
    async with get_session() as session:
        # Получаем все приостановленные задачи
        stmt = select(Task).where(Task.status == "paused")
        result = await session.execute(stmt)
        paused_tasks = result.scalars().all()
        
        resumed_count = 0
        
        for task in paused_tasks:
            try:
                # Возобновляем Celery задачу
                celery_resume_task(task.task_id)
                
                # Обновляем статус в БД
                task.status = "running"
                task.paused_at = None
                resumed_count += 1
                
                # Уведомление через WebSocket
                await notify_task_update(task.to_dict())
                
            except Exception as e:
                logger.error(f"Failed to resume task {task.id}: {e}")
                continue
        
        await session.commit()
        
        return {
            "status": "success",
            "resumed_count": resumed_count,
            "message": f"Возобновлено задач: {resumed_count}"
        }

@api_router.delete("/tasks/completed")
async def clear_completed_tasks():
    """Удаление завершенных задач"""
    async with get_session() as session:
        stmt = select(Task).where(
            Task.status.in_(["completed", "failed", "cancelled"])
        )
        result = await session.execute(stmt)
        completed_tasks = result.scalars().all()
        
        deleted_count = len(completed_tasks)
        
        for task in completed_tasks:
            await session.delete(task)
        
        await session.commit()
        
        return {
            "status": "success",
            "deleted_count": deleted_count
        }

# ==================== Analytics API ====================

@api_router.get("/analytics/stats")
async def get_analytics_stats():
    """Получение статистики для дашборда"""
    async with get_session() as session:
        # Общая статистика
        total_videos = await session.scalar(
            select(func.count(Video.id)).where(Video.parse_status == "completed")
        )
        
        total_channels = await session.scalar(
            select(func.count(Channel.id))
        )
        
        avg_engagement = await session.scalar(
            select(func.avg(Video.engagement_rate)).where(Video.parse_status == "completed")
        )
        
        # Видео за последнюю неделю
        week_ago = datetime.now() - timedelta(days=7)
        videos_this_week = await session.scalar(
            select(func.count(Video.id)).where(
                and_(
                    Video.created_at >= week_ago,
                    Video.parse_status == "completed"
                )
            )
        )
        
        # Топ категория
        top_category_result = await session.execute(
            select(
                Video.video_category,
                func.count(Video.id).label("count")
            ).where(
                Video.video_category.isnot(None)
            ).group_by(
                Video.video_category
            ).order_by(
                desc("count")
            ).limit(1)
        )
        top_category_row = top_category_result.first()
        top_category = top_category_row[0] if top_category_row else "Не определено"
        
        # Топ видео
        top_videos_result = await session.execute(
            select(Video).where(
                Video.parse_status == "completed"
            ).order_by(
                desc(Video.views)
            ).limit(10)
        )
        top_videos = top_videos_result.scalars().all()
        
        return {
            "stats": {
                "totalVideos": total_videos or 0,
                "totalChannels": total_channels or 0,
                "avgEngagement": round(avg_engagement or 0, 2),
                "topCategory": top_category,
                "videosThisWeek": videos_this_week or 0
            },
            "topVideos": [
                {
                    "id": video.id,
                    "title": video.title,
                    "views": video.views,
                    "engagement_rate": round(video.engagement_rate, 2),
                    "channel_name": video.channel_url.split("/")[-1] if video.channel_url else "Unknown",
                    "thumbnail_url": video.thumbnail_url
                }
                for video in top_videos
            ]
        }

# ==================== Search Queries API ====================

@api_router.get("/search-queries")
async def get_search_queries(limit: int = Query(10, ge=1, le=100)):
    """Получение истории поисковых запросов"""
    async with get_session() as session:
        stmt = select(SearchQuery).order_by(
            desc(SearchQuery.created_at)
        ).limit(limit)
        
        result = await session.execute(stmt)
        queries = result.scalars().all()
        
        return [
            {
                "id": query.id,
                "query": query.query,
                "type": query.search_type,
                "results": query.total_results,
                "date": query.created_at.isoformat() if query.created_at else None
            }
            for query in queries
        ]

# ==================== Settings API ====================

@api_router.get("/settings")
async def get_settings():
    """Получение настроек приложения"""
    return {
        "youtubeApiKey": os.getenv("YOUTUBE_API_KEY", ""),
        "databasePath": os.getenv("DATABASE_PATH", "youtube_data.db"),
        "host": os.getenv("HOST", "127.0.0.1"),
        "port": int(os.getenv("PORT", 8000)),
        "debug": os.getenv("DEBUG", "true").lower() == "true",
        "tempDir": os.getenv("TEMP_DIR", "temp"),
        "exportDir": os.getenv("EXPORT_DIR", "exports"),
        "logsDir": os.getenv("LOGS_DIR", "logs"),
        "redisUrl": os.getenv("REDIS_URL", "redis://localhost:6379/0"),
        "maxResultsPerSearch": int(os.getenv("MAX_RESULTS_PER_SEARCH", 500)),
        "apiRequestDelay": int(os.getenv("API_REQUEST_DELAY", 3)),
        "exportPageSize": int(os.getenv("EXPORT_PAGE_SIZE", 1000)),
        "maxExportRows": int(os.getenv("MAX_EXPORT_ROWS", 50000)),
        "enableDeepAnalysis": os.getenv("ENABLE_DEEP_ANALYSIS", "true").lower() == "true",
        "enableWebsocket": os.getenv("ENABLE_WEBSOCKET", "true").lower() == "true",
        "enableAutoRetry": os.getenv("ENABLE_AUTO_RETRY", "true").lower() == "true"
    }

@api_router.post("/settings")
async def save_settings(settings: Dict[str, Any]):
    """Сохранение настроек (заглушка)"""
    # В реальном приложении здесь бы сохранялись настройки в файл или БД
    return {"status": "success", "message": "Настройки сохранены"}

# ==================== Export/Import API ====================

@api_router.get("/export/database")
async def export_database():
    """Экспорт базы данных"""
    import shutil
    
    try:
        # Создание временного архива
        temp_dir = tempfile.mkdtemp()
        archive_path = os.path.join(temp_dir, "youtube_analyzer_backup.zip")
        
        with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Добавление базы данных
            db_path = os.getenv("DATABASE_PATH", "youtube_data.db")
            if os.path.exists(db_path):
                zipf.write(db_path, "youtube_data.db")
            
            # Добавление конфигурации (без секретных данных)
            config_data = {
                "export_date": datetime.now().isoformat(),
                "version": "1.0.0",
                "settings": {
                    "host": os.getenv("HOST", "127.0.0.1"),
                    "port": int(os.getenv("PORT", 8000)),
                    "debug": os.getenv("DEBUG", "true").lower() == "true"
                }
            }
            zipf.writestr("config.json", json.dumps(config_data, indent=2))
        
        return FileResponse(
            archive_path,
            media_type="application/zip",
            filename=f"youtube_analyzer_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
        
    finally:
        # Очистка
        def cleanup():
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        
        import atexit
        atexit.register(cleanup)

@api_router.post("/import/database")
async def import_database(file: UploadFile = File(...)):
    """Импорт базы данных (заглушка)"""
    # В реальном приложении здесь бы происходило восстановление из архива
    return {"status": "success", "message": "База данных импортирована"}

# ==================== Help API ====================

@api_router.get("/help/content")
async def get_help_content():
    """Получение содержимого справки"""
    help_path = "help.html"
    if os.path.exists(help_path):
        with open(help_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return {"content": content}
    else:
        return {"content": "<h1>Справка не найдена</h1>"}
