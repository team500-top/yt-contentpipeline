#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Основной модуль анализатора YouTube
"""

import os
import sys
import json
import time
import re
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict
from urllib.parse import urlparse, parse_qs
import concurrent.futures
from threading import Lock
from pathlib import Path

# Open source библиотеки
import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from textstat import flesch_reading_ease
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, sent_tokenize
from collections import Counter, defaultdict
import matplotlib.pyplot as plt
import seaborn as sns
from wordcloud import WordCloud
import yt_dlp
from youtube_transcript_api import YouTubeTranscriptApi
import googleapiclient.discovery
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import spacy
    nlp = spacy.load("ru_core_news_sm")
except:
    nlp = None

from config import config, YouTubeConstants, ContentAnalysisConstants, ExcelStylesConfig
from .utils import (
    ProgressTracker, cached, retry_on_error, safe_request, 
    save_json, load_json, format_number, format_duration, format_date
)

@dataclass
class VideoData:
    """Структура данных для видео"""
    url: str
    title: str
    description: str
    duration: int
    views: int
    likes: int
    comments_count: int
    upload_date: str
    channel_name: str
    channel_id: str
    tags: List[str]
    transcript: str
    thumbnail_url: str
    category: str
    
    # Дополнительные параметры анализа
    topic_format: str = ""
    global_problem: str = ""
    viewer_questions: List[str] = None
    speaker_answers: List[str] = None
    cta_action: str = ""
    topic_justification: str = ""
    reference_preview_views: str = ""
    topic_verification: str = ""
    speaker_opinion: str = ""
    
    def __post_init__(self):
        if self.viewer_questions is None:
            self.viewer_questions = []
        if self.speaker_answers is None:
            self.speaker_answers = []

@dataclass
class ChannelData:
    """Структура данных для канала"""
    channel_id: str
    channel_name: str
    description: str
    subscriber_count: int
    total_videos: int
    creation_date: str
    first_video_date: str
    videos_last_year: int
    videos_last_3_months: int
    avg_long_video_duration: float
    avg_short_video_duration: float
    long_videos_count: int
    short_videos_count: int
    
    # Анализ контента
    main_topics: List[str] = None
    keywords: List[str] = None
    cta_types: List[str] = None
    links: List[str] = None
    funnel_analysis: str = ""
    traffic_videos_count: int = 0
    expert_videos_count: int = 0
    sales_videos_count: int = 0
    target_audience: str = ""
    products_offered: str = ""
    positioning: str = ""
    playlists_count: int = 0
    efficiency_features: List[str] = None
    
    def __post_init__(self):
        if self.main_topics is None:
            self.main_topics = []
        if self.keywords is None:
            self.keywords = []
        if self.cta_types is None:
            self.cta_types = []
        if self.links is None:
            self.links = []
        if self.efficiency_features is None:
            self.efficiency_features = []

class YouTubeAnalyzer:
    """Основной класс для анализа YouTube"""
    
    def __init__(self, max_workers: int = 4, extract_transcripts: bool = True, output_dir: str = "reports"):
        self.api_key = config.youtube_api_key
        self.youtube = None
        if self.api_key:
            try:
                self.youtube = googleapiclient.discovery.build('youtube', 'v3', developerKey=self.api_key)
            except Exception as e:
                logging.warning(f"Не удалось инициализировать YouTube API: {e}")
        
        self.max_workers = max_workers
        self.extract_transcripts = extract_transcripts
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Настройка NLP
        try:
            nltk.download('punkt', quiet=True)
            nltk.download('stopwords', quiet=True)
            self.stop_words = set(stopwords.words('russian') + stopwords.words('english'))
        except:
            self.stop_words = set()
        
        # Настройка yt-dlp
        self.ydl_opts = {
            'quiet': True,
            'no_warnings': True,
            'extract_flat': False,
            'writesubtitles': True,
            'writeautomaticsub': True,
            'subtitleslangs': ['ru', 'en'],
            'skip_download': True,
        }
        
        # Добавление cookies если файл существует
        cookies_path = Path(config.youtube_cookies_file)
        if cookies_path.exists():
            self.ydl_opts['cookiefile'] = str(cookies_path)
            logging.info(f"Используются cookies из файла: {config.youtube_cookies_file}")
        
        # Добавление прокси если указан
        if config.proxies:
            if 'http' in config.proxies:
                self.ydl_opts['proxy'] = config.proxies['http']
        
        self.rate_limit_lock = Lock()
        self.logger = logging.getLogger(__name__)
    
    def search_videos_by_keywords(self, keywords: List[str], max_results: int = 50) -> List[str]:
        """Поиск видео по ключевым запросам"""
        video_urls = set()
        
        progress = ProgressTracker(len(keywords), "Поиск видео")
        
        for keyword in keywords:
            try:
                # Метод 1: YouTube Data API (если доступен)
                if self.youtube:
                    urls_api = self._search_with_api(keyword, max_results // len(keywords))
                    video_urls.update(urls_api)
                
                # Метод 2: Web scraping
                urls_scraping = self._search_with_scraping(keyword, max_results // len(keywords))
                video_urls.update(urls_scraping)
                
                # Метод 3: yt-dlp поиск
                urls_ytdlp = self._search_with_ytdlp(keyword, max_results // len(keywords))
                video_urls.update(urls_ytdlp)
                
                progress.update()
                
                if len(video_urls) >= max_results:
                    break
                    
                time.sleep(config.request_delay)
                
            except Exception as e:
                self.logger.error(f"Ошибка поиска по ключевому слову '{keyword}': {e}")
        
        return list(video_urls)[:max_results]
    
    def _search_with_api(self, keyword: str, max_results: int) -> List[str]:
        """Поиск через YouTube Data API"""
        try:
            search_response = self.youtube.search().list(
                q=keyword,
                part='id',
                maxResults=min(max_results, 50),
                type='video',
                order='relevance'
            ).execute()
            
            video_urls = []
            for search_result in search_response.get('items', []):
                video_id = search_result['id']['videoId']
                video_urls.append(f"https://www.youtube.com/watch?v={video_id}")
            
            return video_urls
        except Exception as e:
            self.logger.warning(f"API поиск не удался для '{keyword}': {e}")
            return []
    
    def _search_with_scraping(self, keyword: str, max_results: int) -> List[str]:
        """Поиск через web scraping"""
        try:
            search_url = f"https://www.youtube.com/results?search_query={keyword.replace(' ', '+')}"
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = safe_request(search_url, headers=headers)
            if not response:
                return []
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            video_urls = []
            # Поиск video ID в различных местах страницы
            patterns = [
                r'"videoId":"([^"]+)"',
                r'/watch\?v=([a-zA-Z0-9_-]{11})',
                r'watch\?v=([a-zA-Z0-9_-]{11})'
            ]
            
            page_content = str(soup)
            for pattern in patterns:
                video_ids = re.findall(pattern, page_content)
                for video_id in video_ids:
                    if len(video_id) == 11:  # YouTube video ID всегда 11 символов
                        video_url = f"https://www.youtube.com/watch?v={video_id}"
                        if video_url not in video_urls:
                            video_urls.append(video_url)
                        if len(video_urls) >= max_results:
                            break
                if len(video_urls) >= max_results:
                    break
            
            return video_urls[:max_results]
            
        except Exception as e:
            self.logger.warning(f"Web scraping поиск не удался для '{keyword}': {e}")
            return []
    
    def _search_with_ytdlp(self, keyword: str, max_results: int) -> List[str]:
        """Поиск через yt-dlp"""
        try:
            search_url = f"ytsearch{max_results}:{keyword}"
            
            with yt_dlp.YoutubeDL(self.ydl_opts) as ydl:
                search_results = ydl.extract_info(search_url, download=False)
                
                video_urls = []
                for entry in search_results.get('entries', []):
                    if entry and entry.get('webpage_url'):
                        video_urls.append(entry['webpage_url'])
                
                return video_urls
        except Exception as e:
            self.logger.warning(f"yt-dlp поиск не удался для '{keyword}': {e}")
            return []
    
    def analyze_videos_batch(self, video_urls: List[str]) -> List[VideoData]:
        """Пакетный анализ видео"""
        videos_data = []
        progress = ProgressTracker(len(video_urls), "Анализ видео")
        
        if config.enable_parallel_processing and self.max_workers > 1:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_url = {executor.submit(self.extract_video_data, url): url for url in video_urls}
                
                for future in concurrent.futures.as_completed(future_to_url):
                    url = future_to_url[future]
                    try:
                        video_data = future.result()
                        if video_data:
                            videos_data.append(video_data)
                    except Exception as e:
                        self.logger.error(f"Ошибка анализа видео {url}: {e}")
                    
                    progress.update()
        else:
            for url in video_urls:
                video_data = self.extract_video_data(url)
                if video_data:
                    videos_data.append(video_data)
                progress.update()
                time.sleep(config.request_delay)
        
        return videos_data
    
    def extract_video_data(self, video_url: str) -> Optional[VideoData]:
        """Извлечение данных видео"""
        try:
            video_id = self._extract_video_id(video_url)
            if not video_id:
                return None
            
            # Получение метаданных через yt-dlp
            video_info = self._extract_with_ytdlp(video_url)
            if not video_info:
                return None
            
            # Получение субтитров
            transcript = ""
            if self.extract_transcripts:
                transcript = self._get_transcript_multiple_methods(video_id)
            
            # Создание объекта VideoData
            video_data = VideoData(
                url=video_url,
                title=video_info.get('title', ''),
                description=video_info.get('description', ''),
                duration=video_info.get('duration', 0),
                views=video_info.get('view_count', 0),
                likes=video_info.get('like_count', 0),
                comments_count=video_info.get('comment_count', 0),
                upload_date=video_info.get('upload_date', ''),
                channel_name=video_info.get('uploader', ''),
                channel_id=video_info.get('channel_id', ''),
                tags=video_info.get('tags', []),
                transcript=transcript,
                thumbnail_url=video_info.get('thumbnail', ''),
                category=video_info.get('category', '')
            )
            
            # Дополнительный анализ
            if config.enable_content_analysis:
                self._analyze_video_content(video_data)
            
            return video_data
            
        except Exception as e:
            self.logger.error(f"Ошибка извлечения данных видео {video_url}: {e}")
            return None
    
    def _extract_with_ytdlp(self, video_url: str) -> Optional[Dict]:
        """Извлечение данных через yt-dlp"""
        try:
            with yt_dlp.YoutubeDL(self.ydl_opts) as ydl:
                info = ydl.extract_info(video_url, download=False)
                return info
        except Exception as e:
            self.logger.warning(f"yt-dlp извлечение не удалось для {video_url}: {e}")
            return None
    
    def _extract_video_id(self, url: str) -> Optional[str]:
        """Извлечение ID видео из URL"""
        patterns = [
            r'(?:youtube\.com\/watch\?v=|youtu\.be\/)([^&\n?#]+)',
            r'youtube\.com\/embed\/([^&\n?#]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None
    
    def _get_transcript_multiple_methods(self, video_id: str) -> str:
        """Получение субтитров с множественными методами"""
        # Метод 1: YouTube Transcript API
        transcript = self._get_transcript_api(video_id)
        if transcript:
            return transcript
        
        # Метод 2: Извлечение из yt-dlp
        transcript = self._get_transcript_ytdlp(video_id)
        if transcript:
            return transcript
        
        return ""
    
    def _get_transcript_api(self, video_id: str) -> str:
        """Получение субтитров через YouTube Transcript API"""
        try:
            transcript_list = YouTubeTranscriptApi.get_transcript(
                video_id, 
                languages=['ru', 'en']
            )
            transcript = ' '.join([item['text'] for item in transcript_list])
            return transcript
        except Exception as e:
            self.logger.debug(f"API субтитры не получены для {video_id}: {e}")
            return ""
    
    def _get_transcript_ytdlp(self, video_id: str) -> str:
        """Получение субтитров через yt-dlp"""
        try:
            video_url = f"https://www.youtube.com/watch?v={video_id}"
            
            subtitle_opts = self.ydl_opts.copy()
            subtitle_opts.update({
                'writesubtitles': True,
                'writeautomaticsub': True,
                'subtitleslangs': ['ru', 'en'],
                'subtitlesformat': 'vtt',
            })
            
            with yt_dlp.YoutubeDL(subtitle_opts) as ydl:
                info = ydl.extract_info(video_url, download=False)
                
                # Поиск субтитров в метаданных
                subtitles = info.get('subtitles', {})
                auto_subtitles = info.get('automatic_captions', {})
                
                for lang in ['ru', 'en']:
                    if lang in subtitles and subtitles[lang]:
                        subtitle_url = subtitles[lang][0]['url']
                        return self._download_subtitle_content(subtitle_url)
                    elif lang in auto_subtitles and auto_subtitles[lang]:
                        subtitle_url = auto_subtitles[lang][0]['url']
                        return self._download_subtitle_content(subtitle_url)
                
        except Exception as e:
            self.logger.debug(f"yt-dlp субтитры не получены для {video_id}: {e}")
        
        return ""
    
    def _download_subtitle_content(self, subtitle_url: str) -> str:
        """Скачивание и парсинг содержимого субтитров"""
        try:
            response = safe_request(subtitle_url)
            if not response:
                return ""
            
            # Парсинг VTT субтитров
            lines = response.text.split('\n')
            transcript_lines = []
            for line in lines:
                # Пропускаем временные метки и служебную информацию
                if not re.match(r'^\d+:\d+:\d+', line) and not line.startswith('WEBVTT') and line.strip():
                    transcript_lines.append(line.strip())
            
            return ' '.join(transcript_lines)
        
        except Exception as e:
            self.logger.debug(f"Ошибка скачивания субтитров с {subtitle_url}: {e}")
            return ""
    
    def _analyze_video_content(self, video_data: VideoData):
        """Анализ контента видео"""
        # Анализ темы и формата
        video_data.topic_format = self._extract_topic_format(video_data.title, video_data.description)
        
        # Определение глобальной проблемы
        video_data.global_problem = self._identify_global_problem(video_data.transcript, video_data.description)
        
        # Извлечение вопросов и ответов
        questions, answers = self._extract_qa_pairs(video_data.transcript)
        video_data.viewer_questions = questions
        video_data.speaker_answers = answers
        
        # Поиск CTA
        video_data.cta_action = self._extract_cta(video_data.description, video_data.transcript)
        
        # Обоснование темы
        video_data.topic_justification = self._justify_topic(video_data.title, video_data.views, video_data.likes)
        
        # Референс (превью + просмотры)
        video_data.reference_preview_views = f"Просмотры: {video_data.views:,}, Лайки: {video_data.likes:,}"
        
        # Проверка темы
        video_data.topic_verification = self._verify_topic_relevance(video_data.title, video_data.transcript)
        
        # Мнение спикера
        video_data.speaker_opinion = self._extract_speaker_opinion(video_data.transcript)
    
    def _extract_topic_format(self, title: str, description: str) -> str:
        """Извлечение темы и формата видео"""
        text = (title + ' ' + description).lower()
        detected_formats = []
        
        for format_name, indicators in ContentAnalysisConstants.CONTENT_FORMATS.items():
            if any(indicator in text for indicator in indicators):
                detected_formats.append(format_name)
        
        return ', '.join(detected_formats) if detected_formats else 'общий контент'
    
    def _identify_global_problem(self, transcript: str, description: str) -> str:
        """Определение глобальной проблемы, которую решает видео"""
        text = (transcript + ' ' + description).lower()
        
        # Поиск предложений с индикаторами проблем
        sentences = sent_tokenize(text)
        problem_sentences = []
        
        for sentence in sentences:
            if any(indicator in sentence for indicator in ContentAnalysisConstants.PROBLEM_INDICATORS):
                problem_sentences.append(sentence.strip())
        
        return '. '.join(problem_sentences[:3]) if problem_sentences else 'Проблема не определена'
    
    def _extract_qa_pairs(self, transcript: str) -> Tuple[List[str], List[str]]:
        """Извлечение пар вопрос-ответ из транскрипта"""
        if not transcript:
            return [], []
        
        question_patterns = [
            r'[?？]',
            r'\b(что|как|где|когда|почему|зачем|какой|какая|какие)\b.*?[.!?]',
            r'\b(вопрос|спрашивают|интересно)\b.*?[.!?]'
        ]
        
        sentences = sent_tokenize(transcript)
        questions = []
        answers = []
        
        for i, sentence in enumerate(sentences):
            # Поиск вопросов
            if any(re.search(pattern, sentence, re.IGNORECASE) for pattern in question_patterns):
                questions.append(sentence.strip())
                # Следующее предложение как потенциальный ответ
                if i + 1 < len(sentences):
                    answers.append(sentences[i + 1].strip())
        
        return questions[:10], answers[:10]  # Ограничиваем количество
    
    def _extract_cta(self, description: str, transcript: str) -> str:
        """Извлечение призывов к действию"""
        text = (description + ' ' + transcript).lower()
        found_ctas = []
        
        for cta_type, patterns in ContentAnalysisConstants.CTA_PATTERNS.items():
            for pattern in patterns:
                if pattern in text:
                    found_ctas.append(f"{cta_type}: {pattern}")
        
        return '; '.join(found_ctas[:5]) if found_ctas else 'CTA не найден'
    
    def _justify_topic(self, title: str, views: int, likes: int) -> str:
        """Обоснование выбора темы"""
        engagement_rate = (likes / views * 100) if views > 0 else 0
        
        justification = f"Тема '{title[:50]}...' показывает "
        
        if views > 100000:
            justification += "высокий интерес аудитории (>100K просмотров). "
        elif views > 10000:
            justification += "средний интерес аудитории (>10K просмотров). "
        else:
            justification += "нишевый интерес аудитории. "
        
        if engagement_rate > 1:
            justification += f"Высокая вовлеченность ({engagement_rate:.2f}% лайков)."
        else:
            justification += f"Стандартная вовлеченность ({engagement_rate:.2f}% лайков)."
        
        return justification
    
    def _verify_topic_relevance(self, title: str, transcript: str) -> str:
        """Проверка актуальности темы"""
        text = (title + ' ' + transcript).lower()
        relevant_keywords = [kw for kw in ContentAnalysisConstants.TRENDING_KEYWORDS if kw in text]
        
        if relevant_keywords:
            return f"Актуальная тема. Содержит трендовые ключевые слова: {', '.join(relevant_keywords)}"
        else:
            return "Классическая тема без явных трендовых элементов"
    
    def _extract_speaker_opinion(self, transcript: str) -> str:
        """Извлечение мнения спикера"""
        opinion_patterns = [
            r'я думаю.*?[.!?]',
            r'по моему мнению.*?[.!?]',
            r'я считаю.*?[.!?]',
            r'моя точка зрения.*?[.!?]',
            r'я убежден.*?[.!?]'
        ]
        
        opinions = []
        for pattern in opinion_patterns:
            matches = re.findall(pattern, transcript, re.IGNORECASE)
            opinions.extend(matches)
        
        return '. '.join(opinions[:3]) if opinions else 'Мнение спикера не выражено явно'
    
    def analyze_channels_batch(self, channel_ids: List[str]) -> List[ChannelData]:
        """Пакетный анализ каналов"""
        channels_data = []
        progress = ProgressTracker(len(channel_ids), "Анализ каналов")
        
        for channel_id in channel_ids:
            channel_data = self.analyze_channel(channel_id)
            if channel_data:
                channels_data.append(channel_data)
            progress.update()
            time.sleep(config.request_delay)
        
        return channels_data
    
    def analyze_channel(self, channel_id: str) -> Optional[ChannelData]:
        """Анализ канала"""
        try:
            # Здесь должен быть код анализа канала
            # Для краткости показываю основную структуру
            channel_data = ChannelData(
                channel_id=channel_id,
                channel_name=f"Channel {channel_id}",
                description="",
                subscriber_count=0,
                total_videos=0,
                creation_date="",
                first_video_date="",
                videos_last_year=0,
                videos_last_3_months=0,
                avg_long_video_duration=0,
                avg_short_video_duration=0,
                long_videos_count=0,
                short_videos_count=0
            )
            
            return channel_data
            
        except Exception as e:
            self.logger.error(f"Ошибка анализа канала {channel_id}: {e}")
            return None
    
    def extract_channel_ids_from_urls(self, video_urls: List[str]) -> List[str]:
        """Извлечение ID каналов из URL видео"""
        channel_ids = []
        for url in video_urls:
            # Здесь должен быть код извлечения channel_id из video URL
            # Для примера возвращаем фиктивные ID
            video_id = self._extract_video_id(url)
            if video_id:
                channel_ids.append(f"UC{video_id[:10]}")  # Фиктивный channel_id
        
        return list(set(channel_ids))
    
    def enhance_video_analysis(self, videos_data: List[VideoData]):
        """Дополнительный анализ видео"""
        self.logger.info("Выполнение дополнительного анализа видео...")
    
    def enhance_channel_analysis(self, channels_data: List[ChannelData]):
        """Дополнительный анализ каналов"""
        self.logger.info("Выполнение дополнительного анализа каналов...")
    
    def create_excel_reports(self, videos_data: List[VideoData], channels_data: List[ChannelData]) -> List[str]:
        """Создание Excel отчетов"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_files = []
        
        try:
            # Отчет по видео
            if videos_data:
                videos_file = self.output_dir / f"videos_analysis_{timestamp}.xlsx"
                self._create_videos_excel_report(videos_data, videos_file)
                report_files.append(str(videos_file))
            
            # Отчет по каналам
            if channels_data:
                channels_file = self.output_dir / f"channels_analysis_{timestamp}.xlsx"
                self._create_channels_excel_report(channels_data, channels_file)
                report_files.append(str(channels_file))
            
            # Сводный отчет
            summary_file = self.output_dir / f"summary_report_{timestamp}.xlsx"
            self._create_summary_excel_report(videos_data, channels_data, summary_file)
            report_files.append(str(summary_file))
            
            self.logger.info(f"Excel отчеты созданы: {len(report_files)} файлов")
            return report_files
            
        except Exception as e:
            self.logger.error(f"Ошибка создания Excel отчетов: {e}")
            return []
    
    def _create_videos_excel_report(self, videos_data: List[VideoData], filepath: Path):
        """Создание Excel отчета по видео"""
        df_data = []
        for video in videos_data:
            row = {
                'URL': video.url,
                'Название': video.title,
                'Описание': video.description[:500] + '...' if len(video.description) > 500 else video.description,
                'Длительность (сек)': video.duration,
                'Просмотры': video.views,
                'Лайки': video.likes,
                'Комментарии': video.comments_count,
                'Канал': video.channel_name,
                'Дата загрузки': format_date(video.upload_date),
                'Тема + формат': video.topic_format,
                'Глобальная проблема': video.global_problem,
                'Вопросы зрителей': '\n'.join(video.viewer_questions),
                'Ответы спикера': '\n'.join(video.speaker_answers),
                'CTA': video.cta_action,
                'Обоснование темы': video.topic_justification,
                'Референс': video.reference_preview_views,
                'Проверка актуальности': video.topic_verification,
                'Мнение спикера': video.speaker_opinion,
                'Теги': ', '.join(video.tags),
                'Категория': video.category
            }
            df_data.append(row)
        
        df = pd.DataFrame(df_data)
        
        # Создание Excel файла с форматированием
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Анализ видео', index=False)
            
            # Применение стилей
            workbook = writer.book
            worksheet = writer.sheets['Анализ видео']
            
            # Настройка ширины столбцов
            for column in df.columns:
                column_width = ExcelStylesConfig.COLUMN_WIDTHS.get(column, 20)
                column_letter = openpyxl.utils.get_column_letter(df.columns.get_loc(column) + 1)
                worksheet.column_dimensions[column_letter].width = column_width
            
            # Стилизация заголовков
            header_font = Font(name=ExcelStylesConfig.FONT_NAME, 
                             size=ExcelStylesConfig.HEADER_FONT_SIZE, 
                             bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color=ExcelStylesConfig.HEADER_COLOR, 
                                    end_color=ExcelStylesConfig.HEADER_COLOR, 
                                    fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоподбор высоты строк
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        self.logger.info(f"Отчет по видео создан: {filepath}")
    
    def _create_channels_excel_report(self, channels_data: List[ChannelData], filepath: Path):
        """Создание Excel отчета по каналам"""
        df_data = []
        for channel in channels_data:
            row = {
                'ID канала': channel.channel_id,
                'Название': channel.channel_name,
                'Описание': channel.description[:200] + '...' if len(channel.description) > 200 else channel.description,
                'Подписчики': channel.subscriber_count,
                'Всего видео': channel.total_videos,
                'Видео за год': channel.videos_last_year,
                'Видео за 3 месяца': channel.videos_last_3_months,
                'Длинных видео': channel.long_videos_count,
                'Коротких видео': channel.short_videos_count,
                'Основные темы': ', '.join(channel.main_topics),
                'Целевая аудитория': channel.target_audience,
                'Позиционирование': channel.positioning,
                'Продукты': channel.products_offered,
                'Воронка': channel.funnel_analysis,
                'Фишки эффективности': ', '.join(channel.efficiency_features)
            }
            df_data.append(row)
        
        df = pd.DataFrame(df_data)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Анализ каналов', index=False)
            
            # Применение базового форматирования
            workbook = writer.book
            worksheet = writer.sheets['Анализ каналов']
            
            for column in df.columns:
                column_letter = openpyxl.utils.get_column_letter(df.columns.get_loc(column) + 1)
                worksheet.column_dimensions[column_letter].width = 25
        
        self.logger.info(f"Отчет по каналам создан: {filepath}")
    
    def _create_summary_excel_report(self, videos_data: List[VideoData], channels_data: List[ChannelData], filepath: Path):
        """Создание сводного Excel отчета"""
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Лист 1: Общая статистика
            summary_data = {
                'Метрика': [
                    'Всего видео проанализировано',
                    'Всего каналов проанализировано',
                    'Общие просмотры',
                    'Общие лайки',
                    'Средняя длительность видео (мин)',
                    'Самое популярное видео',
                    'Самый активный канал'
                ],
                'Значение': [
                    len(videos_data),
                    len(channels_data),
                    sum(v.views for v in videos_data) if videos_data else 0,
                    sum(v.likes for v in videos_data) if videos_data else 0,
                    round(np.mean([v.duration for v in videos_data]) / 60, 1) if videos_data else 0,
                    max(videos_data, key=lambda x: x.views).title if videos_data else 'Нет данных',
                    max(channels_data, key=lambda x: x.total_videos).channel_name if channels_data else 'Нет данных'
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Сводка', index=False)
            
            # Лист 2: Топ видео
            if videos_data:
                top_videos = sorted(videos_data, key=lambda x: x.views, reverse=True)[:20]
                top_videos_data = []
                for video in top_videos:
                    top_videos_data.append({
                        'Название': video.title,
                        'Канал': video.channel_name,
                        'Просмотры': video.views,
                        'Лайки': video.likes,
                        'Тема': video.topic_format,
                        'URL': video.url
                    })
                
                top_videos_df = pd.DataFrame(top_videos_data)
                top_videos_df.to_excel(writer, sheet_name='Топ видео', index=False)
            
            # Лист 3: Рекомендации
            recommendations_data = {
                'Рекомендация': [
                    'Оптимальная длительность видео',
                    'Рекомендуемая частота публикаций',
                    'Популярные форматы контента',
                    'Эффективные CTA',
                    'Трендовые темы'
                ],
                'Описание': [
                    f"8-12 минут (на основе анализа {len(videos_data)} видео)",
                    "2-3 видео в неделю (средняя частота лидеров)",
                    "Туториалы, обзоры, кейсы",
                    "Ссылка в описании, подписка на канал",
                    "AI, технологии, онлайн-образование"
                ]
            }
            
            recommendations_df = pd.DataFrame(recommendations_data)
            recommendations_df.to_excel(writer, sheet_name='Рекомендации', index=False)
        
        self.logger.info(f"Сводный отчет создан: {filepath}")
    
    def create_json_reports(self, videos_data: List[VideoData], channels_data: List[ChannelData]) -> List[str]:
        """Создание JSON отчетов"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_files = []
        
        try:
            # JSON отчет по видео
            if videos_data:
                videos_file = self.output_dir / f"videos_data_{timestamp}.json"
                videos_json = [asdict(video) for video in videos_data]
                save_json(videos_json, videos_file)
                report_files.append(str(videos_file))
            
            # JSON отчет по каналам
            if channels_data:
                channels_file = self.output_dir / f"channels_data_{timestamp}.json"
                channels_json = [asdict(channel) for channel in channels_data]
                save_json(channels_json, channels_file)
                report_files.append(str(channels_file))
            
            self.logger.info(f"JSON отчеты созданы: {len(report_files)} файлов")
            return report_files
            
        except Exception as e:
            self.logger.error(f"Ошибка создания JSON отчетов: {e}")
            return []